import html
import json
import re
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment


INPUT_JSON = "raw_collected_properties.json"
OUTPUT_CSV = "parsed_hazelview_import.csv"
OUTPUT_XLSX = "parsed_hazelview_import.xlsx"
TEMPLATE_XLSX = "properties_import_final.xlsx"

HEADERS = [
    "post_title",
    "post_content",
    "property_type",
    "property_manager_name",
    "property_manager_phone",
    "property_manager_website",
    "address",
    "province",
    "city_name",
    "suite_types",
    "suite_features",
    "amenities",
    "utilities_included",
    "parking",
]

PROVINCE_MAP = {
    "ON": "Ontario",
    "AB": "Alberta",
    "BC": "British Columbia",
    "MB": "Manitoba",
    "NB": "New Brunswick",
    "NL": "Newfoundland and Labrador",
    "NS": "Nova Scotia",
    "PE": "Prince Edward Island",
    "QC": "Quebec",
    "SK": "Saskatchewan",
}

BEDROOM_MAP = {
    0: "Studio",
    1: "One Bedroom",
    2: "Two Bedroom",
    3: "Three Bedroom",
    4: "Four Bedroom",
}

PLUS_DEN_MAP = {
    1: "One Plus Den",
    2: "Two Plus Den",
    3: "Three Plus Den",
    4: "Four Plus Den",
}

SUITE_FEATURE_PATTERNS = [
    (r"9 foot ceilings?", "9 Foot Ceilings"),
    (r"abundance of natural light", "Abundance of Natural Light"),
    (r"air[- ]?condition(?:er|ing)|air-conditioned suites", "Air Conditioner"),
    (r"alarm system", "Alarm System"),
    (r"balcon(?:y|ies)|private patio|private balcony", "Private Balcony/Patio"),
    (r"cable ready", "Cable Ready"),
    (r"carpeted floors?", "Carpeted Floors"),
    (r"city views?", "City Views"),
    (r"dishwasher", "Dishwasher"),
    (r"dryer in suite|washer in suite|in-suite laundry", "In-Suite Laundry"),
    (r"energy efficient lighting", "Energy Efficient Lighting"),
    (r"ensuite bathroom", "Ensuite Bathroom"),
    (r"floor to ceiling windows", "Floor-to-Ceiling Windows"),
    (r"fridge|refrigerator", "Refrigerator"),
    (r"fully renovated kitchen and bathroom", "Renovated Kitchen and Bathroom"),
    (r"fully renovated suite", "Fully Renovated Suite"),
    (r"furnished suite", "Furnished Suite"),
    (r"hardwood and ceramic tile", "Hardwood and Ceramic Tile"),
    (r"in-suite storage|generous in-suite storage", "In-Suite Storage"),
    (r"individual (?:unit )?thermostats", "Individual Thermostats"),
    (r"laminate countertop", "Laminate Countertops"),
    (r"laminate flooring", "Laminate Flooring"),
    (r"microwave", "Microwave"),
    (r"modern finishes", "Modern Finishes"),
    (r"open concept", "Open Concept Layout"),
    (r"open living spaces", "Open Living Spaces"),
    (r"panoramic views", "Panoramic Views"),
    (r"park views", "Park Views"),
    (r"plank flooring", "Plank Flooring"),
    (r"quartz countertops?", "Quartz Countertops"),
    (r"stainless steel appliances?", "Stainless Steel Appliances"),
    (r"\bstove\b", "Stove"),
    (r"vinyl[- ]plank floors?", "Vinyl Plank Flooring"),
    (r"walk-in closet", "Walk-In Closet"),
]

COMMUNITY_AMENITY_PATTERNS = [
    (r"24/7 emergency service", "24/7 Emergency Service"),
    (r"accessible ramp", "Accessible Ramp"),
    (r"bbq patio", "BBQ Patio"),
    (r"bicycle parking", "Bike Storage"),
    (r"boma best certified", "BOMA BEST Certified"),
    (r"close to major highways", "Close to Major Highways"),
    (r"close to public transit", "Close to Public Transit"),
    (r"convenience store", "Convenience Store"),
    (r"door-to-door mail service", "Door-to-Door Mail Service"),
    (r"exercise room|fitness room|fitness centre|fitness center|gym", "Fitness Facility"),
    (r"keyless entry", "Keyless Entry"),
    (r"on-site professional management", "On-Site Professional Management"),
    (r"on-site staff", "On-Site Staff"),
    (r"parks nearby", "Parks Nearby"),
    (r"pet[- ]friendly", "Pet Friendly"),
    (r"pets allowed", "Pets Allowed"),
    (r"public transit", "Public Transit Nearby"),
    (r"schools nearby", "Schools Nearby"),
    (r"shopping nearby", "Shopping Nearby"),
    (r"video surveillance", "Video Surveillance"),
    (r"walk-up", "Walk-Up"),
]

PARKING_PATTERNS = [
    (r"\bunderground\b", "Underground Parking"),
    (r"\bcovered\b", "Covered Parking"),
    (r"\boutdoor\b", "Outdoor Parking"),
    (r"\bindoor\b", "Indoor Parking"),
    (r"\bvisitor\b", "Visitor Parking"),
    (r"\bstreet\b", "Street Parking"),
    (r"\belectric\b", "Electric Vehicle Parking"),
    (r"\bgarage\b", "Garage Parking"),
]

UTILITY_PATTERNS = [
    (r"\bheat\b", "Heat"),
    (r"\bhot water\b", "Hot Water"),
    (r"\bwater\b", "Water"),
    (r"\bhydro\b", "Hydro"),
    (r"\belectricity\b", "Electricity"),
]


def clean(text: str) -> str:
    return re.sub(r"\s+", " ", str(text)).strip()


def unique_keep_order(items: List[str]) -> List[str]:
    out = []
    seen = set()
    for item in items:
        item = clean(item)
        if not item:
            continue
        key = item.lower()
        if key not in seen:
            seen.add(key)
            out.append(item)
    return out


def detect_provider(prop: Dict) -> str:
    provider = clean(prop.get("provider", ""))
    if provider:
        return provider

    base_url = clean(prop.get("base_url", "")).lower()
    if "hazelviewproperties.com" in base_url:
        return "hazelview"

    return ""


def soup_from_html(html_text: str) -> BeautifulSoup:
    return BeautifulSoup(html_text or "", "lxml")


def page_html(page: Dict) -> str:
    return page.get("html", "") or ""


def strip_html(fragment: str) -> str:
    if not fragment:
        return ""
    return clean(BeautifulSoup(html.unescape(fragment), "lxml").get_text(" ", strip=True))


def normalize_label(text: str) -> str:
    text = clean(text)
    if not text:
        return ""

    text = re.sub(r"([a-z])([A-Z])", r"\1 \2", text)
    text = text.replace("Bbq", "BBQ")
    text = text.replace("Boma", "BOMA")

    words = []
    small = {"and", "or", "of", "the", "to", "for", "in", "on", "with", "via", "a", "an"}
    for index, word in enumerate(text.split()):
        lower = word.lower()
        if re.fullmatch(r"[0-9/+-]+", word):
            words.append(word)
        elif word.upper() in {"BBQ", "BOMA"}:
            words.append(word.upper())
        elif index > 0 and lower in small:
            words.append(lower)
        else:
            words.append(lower.capitalize())

    return " ".join(words)


def extract_webpage_amenity_items(main_page: Dict) -> List[str]:
    soup = soup_from_html(page_html(main_page))
    selectors = [
        "section.property-hero .property-lower .right .amenities ul.building-amenities li .name",
        "section.property-hero .amenities ul.building-amenities li .name",
    ]

    for selector in selectors:
        elements = soup.select(selector)
        if not elements:
            continue
        items = [clean(el.get_text(" ", strip=True)) for el in elements if clean(el.get_text(" ", strip=True))]
        return unique_keep_order(items)

    return []


def extract_utility_labels_from_item(item: str) -> List[str]:
    text = clean(item).lower()
    if not text:
        return []

    found = []
    if re.search(r"\bhot water\b", text):
        found.append("Hot Water")
        text = re.sub(r"\bhot water\b", " ", text)
    if re.search(r"\bheat\b", text):
        found.append("Heat")
    if re.search(r"\bwater\b", text):
        found.append("Water")
    if re.search(r"\bhydro\b", text):
        found.append("Hydro")
    if re.search(r"\belectricity\b", text):
        found.append("Electricity")

    return unique_keep_order(found)


def parse_data_locations(main_page: Dict) -> List[dict]:
    soup = soup_from_html(page_html(main_page))
    for tag in soup.find_all(attrs={"data-locations": True}):
        raw = tag.get("data-locations", "")
        if not raw:
            continue
        for candidate in (raw, html.unescape(raw)):
            try:
                data = json.loads(candidate)
                if isinstance(data, list):
                    return [item for item in data if isinstance(item, dict)]
            except Exception:
                continue
    return []


def extract_payload(main_page: Dict) -> Dict:
    items = parse_data_locations(main_page)
    if not items:
        return {}

    first = items[0]
    data = first.get("data", {})
    return data if isinstance(data, dict) else {}


def canonicalize_item(item: str, patterns: List[Tuple[str, str]]) -> str:
    item = clean(item)
    if not item:
        return ""

    for pattern, label in patterns:
        if re.search(pattern, item, flags=re.I):
            return label

    return normalize_label(item)


def collect_items(payload: Dict) -> Tuple[List[str], List[str], str, str]:
    suite_items: List[str] = []
    property_items: List[str] = []

    for amenity in payload.get("Amenities", []):
        if not isinstance(amenity, dict):
            continue

        name = clean(html.unescape(amenity.get("name", "")))
        category = clean(amenity.get("category", "")).lower()
        if not name:
            continue

        if category == "suite":
            suite_items.append(name)
        elif category == "property":
            property_items.append(name)

    if clean(payload.get("pet_friendly", "")) == "1":
        property_items.append("Pet Friendly")

    overview_text = strip_html(payload.get("building_overview", ""))
    location_text = strip_html(payload.get("location_details", ""))

    return (
        unique_keep_order(suite_items),
        unique_keep_order(property_items),
        overview_text,
        location_text,
    )


def extract_core_fields(payload: Dict, base_url: str, main_page: Dict) -> Dict[str, str]:
    title = clean(payload.get("building_name", "")) or clean(main_page.get("title", ""))
    phone = clean(payload.get("phone", ""))
    website = base_url.rstrip("/")

    street_number = clean(payload.get("street_number", ""))
    street_name = clean(payload.get("street_name", ""))
    city = clean(payload.get("city_name", "")) or clean(payload.get("city", ""))
    province_raw = clean(payload.get("province_code", "")) or clean(payload.get("province", ""))
    province = PROVINCE_MAP.get(province_raw, clean(payload.get("province_name", "")) or province_raw)

    street = clean(f"{street_number} {street_name}")
    address = f"{street}, {city}" if street and city else street

    return {
        "post_title": title,
        "property_manager_phone": phone,
        "property_manager_website": website,
        "address": address,
        "province": province,
        "city_name": city,
    }


def extract_post_content(payload: Dict, main_page: Dict) -> str:
    overview = strip_html(payload.get("building_overview", ""))
    if overview:
        return overview
    return clean(main_page.get("meta_description", ""))


def normalize_suite_type(suite: Dict) -> str:
    bed_raw = clean(suite.get("bed", ""))

    try:
        bed_count = int(float(bed_raw))
    except Exception:
        bed_count = None

    if bed_count is not None:
        if bed_count in BEDROOM_MAP:
            return BEDROOM_MAP[bed_count]

    name = clean(suite.get("type_name", ""))
    if not name:
        return ""

    if re.search(r"\bstudio\b", name, flags=re.I):
        return "Studio"
    if re.search(r"\b1\s*bed(?:room|rooms)?\b|\bone\s*bed(?:room|rooms)?\b", name, flags=re.I):
        return "One Bedroom"
    if re.search(r"\b2\s*bed(?:room|rooms)?\b|\btwo\s*bed(?:room|rooms)?\b", name, flags=re.I):
        return "Two Bedroom"
    if re.search(r"\b3\s*bed(?:room|rooms)?\b|\bthree\s*bed(?:room|rooms)?\b", name, flags=re.I):
        return "Three Bedroom"
    if re.search(r"\b4\s*bed(?:room|rooms)?\b|\bfour\s*bed(?:room|rooms)?\b", name, flags=re.I):
        return "Four Bedroom"

    return normalize_label(name)


def extract_suite_types(payload: Dict) -> List[str]:
    found = []
    for suite in payload.get("suites", []):
        if not isinstance(suite, dict):
            continue
        label = normalize_suite_type(suite)
        if label:
            found.append(label)
    return unique_keep_order(found)


def extract_suite_features(payload: Dict) -> List[str]:
    suite_items, _, overview_text, _ = collect_items(payload)

    found = [canonicalize_item(item, SUITE_FEATURE_PATTERNS) for item in suite_items]
    for pattern, label in SUITE_FEATURE_PATTERNS:
        if overview_text and re.search(pattern, overview_text, flags=re.I):
            found.append(label)

    return unique_keep_order(found)


def split_webpage_amenities(main_page: Dict) -> Tuple[List[str], List[str], List[str]]:
    amenity_items = extract_webpage_amenity_items(main_page)

    amenities: List[str] = []
    utilities: List[str] = []
    parking_specific: List[str] = []
    parking_available = False

    for item in amenity_items:
        utility_matches = extract_utility_labels_from_item(item)
        if utility_matches:
            utilities.extend(utility_matches)
            continue

        specific_parking_matches = [label for pattern, label in PARKING_PATTERNS if re.search(pattern, item, flags=re.I)]
        if specific_parking_matches:
            parking_specific.extend(specific_parking_matches)
            continue

        if re.search(r"\bparking available\b", item, flags=re.I):
            parking_available = True
            continue

        amenities.append(item)

    parking = unique_keep_order(parking_specific)
    if not parking and parking_available:
        parking = ["Parking Available"]

    return unique_keep_order(amenities), unique_keep_order(utilities), parking


def parse_parking_rate_data(payload: Dict) -> List[str]:
    raw = clean(payload.get("parking_type_rates", ""))
    if not raw:
        return []

    try:
        data = json.loads(raw)
    except Exception:
        return []

    found = []
    for key, value in data.items():
        if not isinstance(value, dict):
            continue
        if not value.get("available"):
            continue
        found.append(key)

    return found


def extract_parking_from_payload(payload: Dict) -> List[str]:
    raw_parts = []
    raw_parts.extend(re.split(r"[,/;]", strip_html(payload.get("parking", ""))))
    raw_parts.extend(parse_parking_rate_data(payload))

    if clean(payload.get("visitor_parking", "")) == "1":
        raw_parts.append("visitor")

    found = []
    for item in raw_parts:
        item = clean(item)
        if not item:
            continue
        for pattern, label in PARKING_PATTERNS:
            if re.search(pattern, item, flags=re.I):
                found.append(label)

    return unique_keep_order(found)


def parse_property(prop: Dict) -> Dict[str, str]:
    if detect_provider(prop) != "hazelview":
        raise ValueError("Not a Hazelview property")

    base_url = clean(prop.get("base_url", ""))
    main_page = prop.get("pages", {}).get("main", {})
    payload = extract_payload(main_page)
    if not payload:
        raise ValueError("Hazelview embedded property payload not found")

    core = extract_core_fields(payload, base_url, main_page)
    post_content = extract_post_content(payload, main_page)
    suite_types = extract_suite_types(payload)
    suite_features = extract_suite_features(payload)
    amenities, webpage_utilities, webpage_parking = split_webpage_amenities(main_page)
    utilities = unique_keep_order(webpage_utilities)
    parking = webpage_parking or extract_parking_from_payload(payload)

    if "Vinyl Plank Flooring" in suite_features:
        suite_features = [item for item in suite_features if item != "Plank Flooring"]

    return {
        "post_title": core["post_title"],
        "post_content": post_content,
        "property_type": "Apartment",
        "property_manager_name": "Hazelview Properties",
        "property_manager_phone": core["property_manager_phone"],
        "property_manager_website": core["property_manager_website"],
        "address": core["address"],
        "province": core["province"],
        "city_name": core["city_name"],
        "suite_types": "\n".join(suite_types),
        "suite_features": "\n".join(suite_features),
        "amenities": "\n".join(amenities),
        "utilities_included": "\n".join(utilities),
        "parking": "\n".join(parking),
    }


def write_outputs(rows: List[Dict[str, str]]) -> None:
    df = pd.DataFrame(rows, columns=HEADERS)
    df.to_csv(OUTPUT_CSV, index=False)

    template = Path(TEMPLATE_XLSX)
    if template.exists():
        wb = load_workbook(template)
        ws = wb["Import"]

        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        for r_idx, row in enumerate(rows, start=2):
            for c_idx, header in enumerate(HEADERS, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=row.get(header, ""))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[r_idx].height = 140

        wb.save(OUTPUT_XLSX)
    else:
        df.to_excel(OUTPUT_XLSX, index=False)


def main() -> None:
    raw = json.loads(Path(INPUT_JSON).read_text(encoding="utf-8"))
    rows = []

    for prop in raw:
        if detect_provider(prop) != "hazelview":
            continue

        try:
            row = parse_property(prop)
            rows.append(row)
            print(f"Parsed: {row['post_title']}")
        except Exception as e:
            print(f"FAILED: {prop.get('base_url', '')} -> {e}")

    write_outputs(rows)
    print(f"Saved -> {OUTPUT_CSV}")
    print(f"Saved -> {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
