import json
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


INPUT_JSON = "properties_raw.json"
OUTPUT_XLSX = "properties_import_final.xlsx"

HEADERS = [
    "post_title",
    "post_content",
    "property_type",
    "property_manager_name",
    "property_manager_phone",
    "property_manager_website",
    "address",
    "province",
    "city_name",
    "suite_types",
    "suite_features",
    "amenities",
    "utilities_included",
    "parking",
]

PROVINCE_MAP = {
    "ON": "Ontario",
    "AB": "Alberta",
    "BC": "British Columbia",
    "MB": "Manitoba",
    "NB": "New Brunswick",
    "NL": "Newfoundland and Labrador",
    "NS": "Nova Scotia",
    "PE": "Prince Edward Island",
    "QC": "Quebec",
    "SK": "Saskatchewan",
}

PARKING_PATTERNS = [
    (r"\bindoor and outdoor parking\b", "Indoor and Outdoor Parking"),
    (r"\bindoor/outdoor parking\b", "Indoor and Outdoor Parking"),
    (r"\bindoor parking\b", "Indoor Parking"),
    (r"\boutdoor parking\b", "Outdoor Parking"),
    (r"\bunderground parking\b", "Underground Parking"),
    (r"\bunderground garage\b", "Underground Parking"),
    (r"\bvisitor parking(?: available)?\b", "Visitor Parking"),
    (r"\btenant and visitor parking\b", "Tenant and Visitor Parking"),
    (r"\btenant parking\b", "Tenant Parking"),
    (r"\bguest parking\b", "Guest Parking"),
    (r"\bcovered parking\b", "Covered Parking"),
    (r"\bsurface parking\b", "Surface Parking"),
    (r"\bgarage parking\b", "Garage Parking"),
    (r"\bparking available\b", "Parking Available"),
]

UTILITY_PATTERNS = [
    (r"\bheat included\b", "Heat"),
    (r"\bwater included\b", "Water"),
    (r"\bhot water included\b", "Hot Water"),
    (r"\bhydro included\b", "Hydro"),
    (r"\belectricity included\b", "Electricity"),
    (r"\butilities included\b", "Utilities Included"),
]

SUITE_TYPE_PATTERNS = [
    (r"\bbachelor\b", "Bachelor"),
    (r"\bstudio\b", "Studio"),
    (r"\bone bedroom plus den\b|\b1[\s-]?bedroom plus den\b", "One Bedroom Plus Den"),
    (r"\bone bedroom\b|\b1[\s-]?bedroom\b", "One Bedroom"),
    (r"\btwo bedroom\b|\b2[\s-]?bedroom\b", "Two Bedroom"),
    (r"\bthree bedroom\b|\b3[\s-]?bedroom\b", "Three Bedroom"),
    (r"\bfour bedroom\b|\b4[\s-]?bedroom\b", "Four Bedroom"),
]

FEATURE_PATTERNS = [
    (r"\bopen-concept kitchens?\b", "Open-Concept Kitchens"),
    (r"\bopen-concept modern kitchens?\b", "Open-Concept Modern Kitchens"),
    (r"\bbreakfast bars?\b", "Breakfast Bars"),
    (r"\blaminate countertops?\b", "Laminate Countertops"),
    (r"\bgranite countertops?\b", "Granite Countertops"),
    (r"\bhard surface countertops?\b", "Hard-Surface Countertops"),
    (r"\btiled backsplashes?\b", "Tiled Backsplashes"),
    (r"\bstainless steel appliances?\b", "Stainless Steel Appliances"),
    (r"\bstainless steel refrigerator\b", "Stainless Steel Refrigerator"),
    (r"\bfridge\b", "Fridge"),
    (r"\brefrigerator\b", "Refrigerator"),
    (r"\bstove\b", "Stove"),
    (r"\bmicrowave\b", "Microwave"),
    (r"\bdishwasher\b", "Dishwasher"),
    (r"\bupgraded bathroom fixtures\b", "Upgraded Bathroom Fixtures"),
    (r"\bnew mirrors\b", "New Mirrors"),
    (r"\bnew vanities\b", "New Vanities"),
    (r"\bfreshly painted units?\b", "Freshly Painted Units"),
    (r"\bnewly finished flooring\b", "Newly Finished Flooring"),
    (r"\bnew flooring\b", "New Flooring"),
    (r"\bnew light fixtures\b", "New Light Fixtures"),
    (r"\bnew hardware fixtures\b", "New Hardware Fixtures"),
    (r"\bhardwood floors?\b", "Hardwood Floors"),
    (r"\bhardwood flooring\b", "Hardwood Flooring"),
    (r"\bair conditioning\b", "Air Conditioning"),
    (r"\bprivate balconies?\b", "Private Balconies"),
    (r"\bbalcon(?:y|ies)\b", "Balconies"),
    (r"\bmodern cabinetry(?: and backsplashes)?\b", "Modern Cabinetry and Backsplashes"),
    (r"\bspacious units?\b", "Spacious Units"),
    (r"\bspacious suites?\b", "Spacious Suites"),
    (r"\bupdated kitchens?(?: and bathrooms?)?\b", "Updated Kitchens and Bathrooms"),
    (r"\brenovated kitchen(?:s)?(?: and bathrooms?)?\b", "Renovated Kitchens and Bathrooms"),
    (r"\bgleaming hardwood floors\b", "Hardwood Floors"),
    (r"\bupgraded lighting\b", "Upgraded Lighting"),
    (r"\bcontemporary cabinetry\b", "Contemporary Cabinetry"),
    (r"\badded in-suite storage\b", "Added In-Suite Storage"),
    (r"\bindividually controlled thermostats\b", "Individually Controlled Thermostats"),
    (r"\blarge,? brightly lit windows\b", "Large, Brightly Lit Windows"),
    (r"\bmirrored sliding doors\b", "Mirrored Sliding Doors"),
    (r"\bdark-stained floors\b", "Dark-Stained Floors"),
]

AMENITY_PATTERNS = [
    (r"\bfitness room\b", "Fitness Room"),
    (r"\bwell-equipped fitness facility\b", "Well-Equipped Fitness Facility"),
    (r"\bfitness facility\b", "Fitness Facility"),
    (r"\bsauna\b", "Sauna"),
    (r"\bparty(?: and| &)games room\b", "Party & Games Room"),
    (r"\bparty room\b", "Party Room"),
    (r"\bgames room\b", "Games Room"),
    (r"\blaundry facilities\b", "Laundry Facilities"),
    (r"\bon-site laundry facility\b", "On-Site Laundry Facilities"),
    (r"\bon-site laundry facilities\b", "On-Site Laundry Facilities"),
    (r"\blaundry room on every floor\b", "Laundry Room on Every Floor"),
    (r"\bkeyless entry\b", "Keyless Entry"),
    (r"\bkeyless fob entry\b", "Keyless Fob Entry"),
    (r"\bon-site management\b", "On-Site Management"),
    (r"\bprofessional on-site management(?: team)?\b", "Professional On-Site Management"),
    (r"\bon-site maintenance staff\b", "On-Site Maintenance Staff"),
    (r"\bfriendly and professional property management\b", "Friendly and Professional Property Management"),
    (r"\bmycommunity portal\b", "MyCommunity Portal"),
    (r"\bsecure building access\b", "Secure Building Access"),
    (r"\bcontrolled access and security cameras\b", "Controlled Access and Security Cameras"),
    (r"\bvideo surveillance\b", "Video Surveillance"),
    (r"\belevators?\b", "Elevators"),
    (r"\bcourtyard\b", "Courtyard"),
    (r"\bbike racks?\b", "Bike Racks"),
    (r"\bmicro farming\b", "Micro Farming"),
    (r"\bstorage lockers?\b", "Storage Lockers"),
    (r"\bstorage lockers available\b", "Storage Lockers Available"),
    (r"\bparcel room\b", "Parcel Room"),
    (r"\bsecure parcel room(?: with tenant access only)?\b", "Secure Parcel Room"),
    (r"\bvisitor parking\b", "Visitor Parking"),
    (r"\boutdoor parking\b", "Outdoor Parking"),
    (r"\bindoor parking\b", "Indoor Parking"),
    (r"\bunderground parking\b", "Underground Parking"),
    (r"\btenant parking\b", "Tenant Parking"),
    (r"\bparks nearby\b", "Parks Nearby"),
    (r"\bschools nearby\b", "Schools Nearby"),
    (r"\bshopping nearby\b", "Shopping Nearby"),
    (r"\bsite superintendent(?:\(s\))?\b", "Site Superintendent(s)"),
    (r"\bpet-friendly\b", "Pet-Friendly"),
    (r"\bpets allowed\b", "Pets Allowed"),
    (r"\binternet ready\b", "Internet Ready"),
    (r"\bcable ready\b", "Cable Ready"),
    (r"\bdoor-to-door mail delivery via canada post\b", "Door-to-Door Mail Delivery via Canada Post"),
    (r"\bvirtual concierge\b", "Virtual Concierge"),
    (r"\bron-site staff\b", "On-Site Staff"),
    (r"\bvisitor parking\b", "Visitor Parking"),
    (r"\bconvenient online payments(?: &| and) maintenance requests\b", "Convenient Online Payments & Maintenance Requests"),
    (r"\bonline payments and communications\b", "Online Payments and Communications"),
    (r"\brent-controlled living\b", "Rent-Controlled Living"),
]

JUNK_EXACT = {
    "HOME", "SUITES", "AMENITIES", "GALLERY", "LOCATION", "RESIDENT PORTAL", "CONTACT",
    "Amenities", "Additional Amenities", "Apartment Benefits:", "Community Benefits:",
    "Apartment Amenities:", "Parking:", "BOOK NOW", "BOOK A TOUR",
    "GET IN TOUCH TODAY", "Get In Touch Today", "Managed by:", "Disclaimer",
    "Share by:", "Available Suites", "Apartment Availability", "Bedrooms",
    "Bathrooms", "Den", "Half Bath", "Rent:", "Area:", "Address", "Button",
    "Want to Learn more?", "Want to learn more?"
}

JUNK_PREFIXES = (
    "All Rights Reserved",
    "Powered by",
    "Our laundry room is available",
    "Our upgraded laundry facilities",
    "Our staff is trained",
    "Stay active with",
    "Relax and unwind",
    "Spend time with",
    "Secure and convenient",
    "From the moment you walk in",
    "MyCommunity Portal is available",
    "On-site maintenance staff to ensure",
    "Keep your vehicle safe and secure",
    "A new level of luxury awaits",
)

PHONE_RE = re.compile(r"^\d{3}-\d{3}-\d{4}$")
POSTAL_RE = r"[A-Z]\d[A-Z]\s?\d[A-Z]\d"


def clean_space(text: str) -> str:
    return re.sub(r"\s+", " ", str(text)).strip()


def unique_keep_order(items):
    out = []
    seen = set()
    for item in items:
        item = clean_space(item)
        if not item:
            continue
        key = item.lower()
        if key not in seen:
            seen.add(key)
            out.append(item)
    return out


def title_case_phrase(s: str) -> str:
    s = clean_space(s)
    if not s:
        return ""
    small = {"and", "or", "of", "the", "to", "for", "in", "on", "with", "via", "from", "at", "a", "an"}
    out = []
    for i, w in enumerate(s.split()):
        lw = w.lower()
        if lw in {"ttc", "on"}:
            out.append(lw.upper())
        elif lw == "mycommunity":
            out.append("MyCommunity")
        elif lw == "on-site":
            out.append("On-Site")
        elif lw == "in-suite":
            out.append("In-Suite")
        elif lw == "pet-friendly":
            out.append("Pet-Friendly")
        elif re.fullmatch(r"[0-9&/().,+\-*]+", w):
            out.append(w)
        elif i > 0 and lw in small:
            out.append(lw)
        else:
            out.append(lw.capitalize())
    s = " ".join(out)
    s = s.replace("Undergound", "Underground")
    s = s.replace("Appliancess", "Appliances")
    s = s.replace("Hard Surface", "Hard-Surface")
    s = s.replace("Door-to-door", "Door-to-Door")
    return s


def sentence_style(s: str) -> str:
    s = clean_space(s)
    if not s:
        return ""
    s = re.sub(r"\s+([,.;:!?])", r"\1", s)
    s = re.sub(r"([,.;:!?])([A-Za-z])", r"\1 \2", s)
    return s[0].upper() + s[1:]


def split_text(text):
    return [clean_space(x) for x in str(text).splitlines() if clean_space(x)]


def all_page_text(prop):
    return "\n".join(page.get("text", "") for page in prop.get("pages", {}).values())


def is_junk(line: str) -> bool:
    if not line:
        return True
    if line in JUNK_EXACT:
        return True
    if PHONE_RE.match(line):
        return True
    if any(line.startswith(prefix) for prefix in JUNK_PREFIXES):
        return True
    if line.isupper() and len(line.split()) <= 4:
        return True
    return False


def normalize_suite_type_text(text: str) -> str:
    text = text.lower()
    replacements = [
        (r"\bone and two-bedroom\b", "one bedroom two bedroom"),
        (r"\btwo and three-bedroom\b", "two bedroom three bedroom"),
        (r"\bthree and four-bedroom\b", "three bedroom four bedroom"),
        (r"\b1 and 2-bedroom\b", "1 bedroom 2 bedroom"),
        (r"\b2 and 3-bedroom\b", "2 bedroom 3 bedroom"),
        (r"\b3 and 4-bedroom\b", "3 bedroom 4 bedroom"),
        (r"\b1, 2,? ?& 3-bedroom\b", "1 bedroom 2 bedroom 3 bedroom"),
        (r"\b1, 2,? and 3-bedroom\b", "1 bedroom 2 bedroom 3 bedroom"),
        (r"\bstudio, 1, 2,? ?& 3-bedroom\b", "studio 1 bedroom 2 bedroom 3 bedroom"),
        (r"\bbachelor, 1 and 2-bedroom\b", "bachelor 1 bedroom 2 bedroom"),
    ]
    for pattern, replacement in replacements:
        text = re.sub(pattern, replacement, text, flags=re.I)
    return text


def parse_manager_name(prop):
    sf = prop.get("scraped_fields", {})
    existing = clean_space(sf.get("property_manager_name", ""))
    if existing:
        return title_case_phrase(existing)

    blob = "\n".join(
        page.get("html", "") + "\n" + page.get("text", "")
        for page in prop.get("pages", {}).values()
    ).lower()

    if "sterling" in blob or "karamar" in blob:
        return "Sterling Karamar"
    if "hazelview" in blob:
        return "Hazelview Properties"
    if "cogir" in blob:
        return "Cogir"

    return ""


def parse_address_fields(prop):
    sf = prop.get("scraped_fields", {})
    address = clean_space(sf.get("address", ""))
    province = clean_space(sf.get("province", ""))
    city = clean_space(sf.get("city_name", ""))

    text = all_page_text(prop)

    if not address:
        m = re.search(
            rf'(\d+[^\n,]*?(?:Street|St\.|Road|Rd\.|Drive|Dr\.|Avenue|Ave\.|Crescent|Court|Boulevard|Blvd\.|Terrace|Way)),?\s*([A-Za-z .\'-]+),\s*(ON|AB|BC|MB|NB|NL|NS|PE|QC|SK)\,?\s*({POSTAL_RE})?',
            text,
            re.I
        )
        if m:
            address = f"{clean_space(m.group(1))}, {title_case_phrase(m.group(2))}"
            province = PROVINCE_MAP[m.group(3).upper()]
            city = title_case_phrase(m.group(2))

    return address, title_case_phrase(province), title_case_phrase(city)


def parse_post_content(prop):
    sf = prop.get("scraped_fields", {})
    home = prop.get("pages", {}).get("home", {})
    home_text = home.get("text", "")
    home_meta = clean_space(home.get("meta_description", ""))

    # Use main page description first
    if home_meta and len(home_meta) > 60:
        return sentence_style(home_meta)

    lines = split_text(home_text)

    about_markers = [
        "About Our Suites",
        "Our Suites",
        f"About {clean_space(sf.get('post_title', ''))}",
    ]

    for marker in about_markers:
        if not marker.strip():
            continue
        capture = False
        collected = []
        for line in lines:
            if line == marker:
                capture = True
                continue
            if capture and line in {
                "LEARN MORE", "Our Amenities", "EXPLORE LOCATION", "MyCommunity Portal", "BOOK A TOUR"
            }:
                break
            if capture and len(line) > 70 and not is_junk(line):
                collected.append(line)
        if collected:
            return sentence_style(" ".join(collected[:2]))

    # fallback only if main page didn't have enough
    sf_post = clean_space(sf.get("post_content", ""))
    if sf_post and len(sf_post) > 60 and not re.search(POSTAL_RE, sf_post):
        return sentence_style(sf_post)

    location_meta = clean_space(prop.get("pages", {}).get("location", {}).get("meta_description", ""))
    if location_meta and len(location_meta) > 60:
        return sentence_style(location_meta)

    return ""


def parse_suite_types(prop):
    sf = prop.get("scraped_fields", {})
    vals = sf.get("suite_types", []) or []
    if vals:
        return "\n".join(unique_keep_order(vals))

    text = normalize_suite_type_text(all_page_text(prop))
    found = []
    for pattern, label in SUITE_TYPE_PATTERNS:
        if re.search(pattern, text, re.I):
            found.append(label)

    return "\n".join(unique_keep_order(found))


def parse_suite_features(prop):
    sf = prop.get("scraped_fields", {})
    existing = sf.get("suite_features", []) or []
    found = list(existing)

    text = "\n".join([
        prop.get("pages", {}).get("home", {}).get("text", ""),
        prop.get("pages", {}).get("suites", {}).get("text", ""),
        prop.get("pages", {}).get("amenities", {}).get("text", ""),
    ])

    for pattern, label in FEATURE_PATTERNS:
        if re.search(pattern, text, re.I):
            found.append(label)

    found = unique_keep_order(found)

    # normalize duplicates / weaker variants
    if "Stainless Steel Appliances" in found:
        found = [x for x in found if x not in {"Stainless Steel Refrigerator", "Fridge", "Refrigerator", "Stove", "Microwave"}]
    if "Private Balconies" in found:
        found = [x for x in found if x != "Balconies"]
    if "Updated Kitchens and Bathrooms" in found and "Renovated Kitchens and Bathrooms" in found:
        found = [x for x in found if x != "Updated Kitchens and Bathrooms"]

    return "\n".join(found)


def parse_parking(prop):
    sf = prop.get("scraped_fields", {})
    found = list(sf.get("parking", []) or [])

    text = all_page_text(prop)

    for pattern, label in PARKING_PATTERNS:
        if re.search(pattern, text, re.I):
            found.append(label)

    found = unique_keep_order(found)

    if "Indoor and Outdoor Parking" in found:
        found = [x for x in found if x not in {"Indoor Parking", "Outdoor Parking"}]

    specific = [x for x in found if x != "Parking Available"]

    if specific:
        return "\n".join(specific)

    if "parking available" in text.lower():
        return "Parking Available"

    return ""


def parse_utilities(prop):
    sf = prop.get("scraped_fields", {})
    found = list(sf.get("utilities_included", []) or [])

    text = all_page_text(prop)

    for pattern, label in UTILITY_PATTERNS:
        if re.search(pattern, text, re.I):
            found.append(label)

    found = unique_keep_order(found)

    if "Utilities Included" in found and len(found) > 1:
        found = [x for x in found if x != "Utilities Included"]

    return "\n".join(found)


def parse_amenities(prop, suite_features_str, parking_str):
    sf = prop.get("scraped_fields", {})
    found = list(sf.get("amenities", []) or [])

    text = "\n".join([
        prop.get("pages", {}).get("amenities", {}).get("text", ""),
        prop.get("pages", {}).get("home", {}).get("text", ""),
    ])

    for pattern, label in AMENITY_PATTERNS:
        if re.search(pattern, text, re.I):
            found.append(label)

    found = unique_keep_order(found)

    suite_feature_set = set(suite_features_str.split("\n")) if suite_features_str else set()
    parking_set = set(parking_str.split("\n")) if parking_str else set()

    cleaned = [x for x in found if x not in suite_feature_set and x not in parking_set]

    return "\n".join(cleaned)


def build_rows(data):
    rows = []

    for prop in data:
        sf = prop.get("scraped_fields", {})
        address, province, city = parse_address_fields(prop)
        suite_features = parse_suite_features(prop)
        parking = parse_parking(prop)

        row = {
            "post_title": title_case_phrase(
                sf.get("post_title") or prop.get("pages", {}).get("home", {}).get("title", "")
            ),
            "post_content": parse_post_content(prop),
            "property_type": title_case_phrase(sf.get("property_type", "Apartment") or "Apartment"),
            "property_manager_name": parse_manager_name(prop),
            "property_manager_phone": clean_space(sf.get("property_manager_phone", "")),
            "property_manager_website": clean_space(sf.get("property_manager_website") or prop.get("url", "")),
            "address": address,
            "province": province,
            "city_name": city,
            "suite_types": parse_suite_types(prop),
            "suite_features": suite_features,
            "amenities": parse_amenities(prop, suite_features, parking),
            "utilities_included": parse_utilities(prop),
            "parking": parking,
        }
        rows.append(row)

    return rows


def write_xlsx(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Import"

    ws.append(HEADERS)
    for row in rows:
        ws.append([row.get(h, "") for h in HEADERS])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 28, "B": 60, "C": 14, "D": 24, "E": 18, "F": 36, "G": 30,
        "H": 14, "I": 16, "J": 22, "K": 34, "L": 38, "M": 24, "N": 24
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 125

    wb.save(output_path)


def main():
    input_path = Path(INPUT_JSON)
    output_path = Path(OUTPUT_XLSX)

    with input_path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    rows = build_rows(data)
    write_xlsx(rows, output_path)

    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()