import json
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment


INPUT_JSON = "raw_collected_properties.json"
OUTPUT_CSV = "parsed_realstar_import.csv"
OUTPUT_XLSX = "parsed_realstar_import.xlsx"
TEMPLATE_XLSX = "properties_import_final.xlsx"

HEADERS = [
    "post_title",
    "post_content",
    "property_type",
    "property_manager_name",
    "property_manager_phone",
    "property_manager_website",
    "address",
    "province",
    "city_name",
    "suite_types",
    "suite_features",
    "amenities",
    "utilities_included",
    "parking",
]

PROVINCE_MAP = {
    "ON": "Ontario",
    "AB": "Alberta",
    "BC": "British Columbia",
    "MB": "Manitoba",
    "NB": "New Brunswick",
    "NL": "Newfoundland and Labrador",
    "NS": "Nova Scotia",
    "PE": "Prince Edward Island",
    "QC": "Quebec",
    "SK": "Saskatchewan",
}

BEDROOM_MAP = {
    0: "",
    1: "One Bedroom",
    2: "Two Bedroom",
    3: "Three Bedroom",
    4: "Four Bedroom",
}

PLUS_DEN_PATTERNS = [
    (r"\b1\s*bed\s*\+\s*den\b|\b1\s*bedroom\s*\+\s*den\b|\bone plus den\b|\bone bedroom plus den\b", "One Plus Den"),
    (r"\b2\s*bed\s*\+\s*den\b|\b2\s*bedroom\s*\+\s*den\b|\btwo plus den\b|\btwo bedroom plus den\b", "Two Plus Den"),
    (r"\b3\s*bed\s*\+\s*den\b|\bthree plus den\b|\bthree bedroom plus den\b", "Three Plus Den"),
]

SUITE_FEATURE_PATTERNS = [
    (r"stainless steel appliances?", "Stainless Steel Appliances"),
    (r"\bdishwasher\b", "Dishwasher"),
    (r"\brefrigerator\b", "Refrigerator"),
    (r"over[- ]the[- ]range microwave|\bmicrowave\b", "Over-the-range Microwave"),
    (r"\bstove\b", "Stove"),
    (r"in[- ]suite washer and dryer|washer and dryer", "In-suite Washer and Dryer"),
    (r"in[- ]suite laundry", "In-suite Laundry"),
    (r"private balcony/?patio|private balcony or patio|private terrace|private balcony|private patio", "Private Balcony/Patio"),
    (r"luxury vinyl plank flooring", "Luxury Vinyl Plank Flooring"),
    (r"hard surface flooring", "Hard Surface Flooring"),
    (r"\bhardwood\b", "Hardwood Flooring"),
    (r"\bceramic\b", "Ceramic Flooring"),
    (r"window coverings|custom roller blinds", "Window Coverings"),
    (r"designer cabinetry|cabinetry", "Designer Cabinetry"),
    (r"quartz countertops?", "Quartz Countertops"),
    (r"recessed lighting", "Recessed Lighting"),
    (r"freshly painted", "Freshly Painted Units"),
    (r"energy efficient appliances?", "Energy Efficient Appliances"),
    (r"air conditioner|air conditioning", "Air Conditioner"),
    (r"water views", "Water Views"),
    (r"refinished kitchen|refinished bathroom", "Refinished Kitchen and Bathroom"),
]

COMMUNITY_AMENITY_PATTERNS = [
    (r"bbq area|outdoor bbq|patio with bbq", "BBQ Area"),
    (r"bike storage", "Bike Storage"),
    (r"\belevators?\b", "Elevators"),
    (r"fitness facility|fitness centre|fitness center|fitness room|\bgym\b", "Fitness Facility"),
    (r"resident lounge|lounge|social room|media room", "Resident Lounge"),
    (r"availability 24 hours", "Availability 24 Hours"),
    (r"on-site management", "On-Site Management"),
    (r"pet friendly", "Pet Friendly"),
    (r"secure entry|controlled access", "Secure Entry"),
    (r"smoke[- ]free", "Smoke-Free Living"),
    (r"laundry facilities|luxury laundry facilities|on-site laundry", "Laundry Facilities"),
    (r"security cameras?", "Security Cameras"),
]

PARKING_PATTERNS = [
    (r"underground parking", "Underground Parking"),
    (r"covered parking", "Covered Parking"),
    (r"surface parking", "Surface Parking"),
    (r"indoor parking", "Indoor Parking"),
    (r"outdoor parking", "Outdoor Parking"),
    (r"garage parking", "Garage Parking"),
]

INCLUDED_UTILITY_PATTERNS = [
    (r"heat included|heat is included|heat included in rent", "Heat"),
    (r"water included|water is included|water included in rent", "Water"),
    (r"hydro included|hydro is included", "Hydro"),
    (r"electricity included|electricity is included", "Electricity"),
    (r"hot water included|hot water is included", "Hot Water"),
]


def clean(text: str) -> str:
    return re.sub(r"\s+", " ", str(text)).strip()


def unique_keep_order(items: List[str]) -> List[str]:
    out = []
    seen = set()
    for item in items:
        item = clean(item)
        if not item:
            continue
        key = item.lower()
        if key not in seen:
            seen.add(key)
            out.append(item)
    return out


def soup_from_html(html: str) -> BeautifulSoup:
    return BeautifulSoup(html or "", "lxml")


def page_text(page: Dict) -> str:
    return page.get("text", "") or ""


def page_html(page: Dict) -> str:
    return page.get("html", "") or ""


def page_json_ld(page: Dict) -> List[dict]:
    data = page.get("json_ld", [])
    return data if isinstance(data, list) else []


def first_apartment_json_ld(items: List[dict]) -> dict:
    for item in items:
        item_type = item.get("@type", [])
        if isinstance(item_type, str):
            item_type = [item_type]
        if "ApartmentComplex" in item_type or "LocalBusiness" in item_type:
            return item
    return {}


def extract_core_fields(main_page: Dict, base_url: str) -> Dict[str, str]:
    item = first_apartment_json_ld(page_json_ld(main_page))

    title = clean(item.get("name", ""))
    phone = clean(item.get("telephone", ""))
    website = clean(item.get("url", "")) or base_url.rstrip("/")

    address_obj = item.get("address", {}) if isinstance(item, dict) else {}
    street = clean(address_obj.get("streetAddress", ""))
    city = clean(address_obj.get("addressLocality", ""))
    province_raw = clean(address_obj.get("addressRegion", ""))
    province = PROVINCE_MAP.get(province_raw, province_raw)

    address = f"{street}, {city}" if street and city else street

    return {
        "post_title": title,
        "property_manager_phone": phone,
        "property_manager_website": website,
        "address": address,
        "province": province,
        "city_name": city,
    }


def extract_post_content(main_page: Dict) -> str:
    text = page_text(main_page)

    # Prefer a bounded "Our suites" section
    patterns = [
        r"Our suites\s+(.*?)(?:AMENITIES|SEE ALL AMENITIES|Property Highlights|Apartment Neighbourhood|$)",
        r"Apartment Amenities\s+(.*?)(?:Community Amenities|Utilities|$)",
    ]
    for pattern in patterns:
        m = re.search(pattern, text, flags=re.I | re.S)
        if m:
            return clean(m.group(1))

    # Fallback to meta description if section is missing
    return clean(main_page.get("meta_description", ""))


def extract_suite_types(floorplans_page: Dict, main_page: Dict) -> List[str]:
    found: List[str] = []

    # 1) JSON-LD floor plans first
    for item in page_json_ld(floorplans_page):
        plans = item.get("accommodationFloorPlan")
        if isinstance(plans, list):
            for plan in plans:
                if not isinstance(plan, dict):
                    continue
                bedrooms = int(plan.get("numberOfBedrooms", 0) or 0)
                label = BEDROOM_MAP.get(bedrooms, "")
                if label:
                    found.append(label)

                name = clean(plan.get("name", ""))
                for pattern, plus_den_label in PLUS_DEN_PATTERNS:
                    if re.search(pattern, name, flags=re.I):
                        found.append(plus_den_label)

    # 2) Text fallback from floorplans and main page
    text = page_text(floorplans_page) + "\n" + page_text(main_page)

    plain_patterns = [
        (r"\bone bedroom\b|\b1 bed\b|\b1 bedroom\b", "One Bedroom"),
        (r"\btwo bedroom\b|\b2 bed\b|\b2 bedroom\b", "Two Bedroom"),
        (r"\bthree bedroom\b|\b3 bed\b|\b3 bedroom\b", "Three Bedroom"),
        (r"\bfour bedroom\b|\b4 bed\b|\b4 bedroom\b", "Four Bedroom"),
    ]
    for pattern, label in plain_patterns:
        if re.search(pattern, text, flags=re.I):
            found.append(label)

    for pattern, label in PLUS_DEN_PATTERNS:
        if re.search(pattern, text, flags=re.I):
            found.append(label)

    return unique_keep_order(found)


def get_section_text_by_heading(soup: BeautifulSoup, heading_keywords: List[str]) -> str:
    """
    Looks for a heading and gathers nearby text from the closest reasonable container.
    This is not perfect, but much better than one giant page-wide regex.
    """
    headings = soup.find_all(re.compile(r"^h[1-6]$"))
    for heading in headings:
        heading_text = clean(heading.get_text(" ", strip=True)).lower()
        if any(keyword.lower() in heading_text for keyword in heading_keywords):
            # Try nearest meaningful container
            container = heading.parent
            for _ in range(4):
                if not container:
                    break
                text = clean(container.get_text(" ", strip=True))
                if len(text) > len(heading_text) + 20:
                    return text
                container = container.parent
    return ""


def extract_selenium_texts(soup: BeautifulSoup, prefixes: List[str]) -> List[str]:
    found = []
    for prefix in prefixes:
        selector = f'[data-selenium-id^="{prefix}"]'
        for el in soup.select(selector):
            text = clean(el.get_text(" ", strip=True))
            if text and not text.lower().startswith("more info"):
                found.append(text)
    return unique_keep_order(found)


def extract_apartment_and_community_blocks(amenities_page: Dict) -> Tuple[List[str], List[str], List[str]]:
    soup = soup_from_html(page_html(amenities_page))

    # Strongest source first: selenium-id blocks if present
    apartment_items = extract_selenium_texts(soup, ["AptAmenity", "ApartmentAmenity"])
    community_items = extract_selenium_texts(soup, ["CommAmenity", "CommunityAmenity"])
    utility_items = extract_selenium_texts(soup, ["utilitiesAmenity", "UtilityAmenity"])

    # Fallback to heading-based section scraping
    if not apartment_items:
        apartment_block = get_section_text_by_heading(soup, ["Apartment Amenities"])
        if apartment_block:
            apartment_items = split_possible_list(apartment_block, drop_headers=["Apartment Amenities"])

    if not community_items:
        community_block = get_section_text_by_heading(soup, ["Community Amenities"])
        if community_block:
            community_items = split_possible_list(community_block, drop_headers=["Community Amenities"])

    if not utility_items:
        utility_block = get_section_text_by_heading(soup, ["Utilities"])
        if utility_block:
            utility_items = split_possible_list(utility_block, drop_headers=["Utilities"])

    return apartment_items, community_items, utility_items


def split_possible_list(text: str, drop_headers: Optional[List[str]] = None) -> List[str]:
    text = clean(text)
    if not text:
        return []

    drop_headers = drop_headers or []
    for header in drop_headers:
        text = re.sub(rf"^{re.escape(header)}\s*", "", text, flags=re.I)

    # Break apart likely list items
    parts = re.split(
        r"\s{2,}|(?<=[a-z0-9\)]) (?=[A-Z][a-z])|•|\*",
        text,
    )

    cleaned = []
    for part in parts:
        part = clean(part)
        if not part:
            continue
        if drop_headers and any(part.lower() == h.lower() for h in drop_headers):
            continue
        cleaned.append(part)

    return unique_keep_order(cleaned)


def extract_suite_features(main_page: Dict, amenities_page: Dict) -> List[str]:
    apartment_items, _, _ = extract_apartment_and_community_blocks(amenities_page)

    texts = [
        extract_post_content(main_page),
        " ".join(apartment_items),
    ]

    found = []
    for text in texts:
        for pattern, label in SUITE_FEATURE_PATTERNS:
            if re.search(pattern, text, flags=re.I):
                found.append(label)

    return unique_keep_order(found)


def extract_community_amenities(main_page: Dict, amenities_page: Dict) -> List[str]:
    _, community_items, _ = extract_apartment_and_community_blocks(amenities_page)

    texts = [
        " ".join(community_items),
        page_text(main_page),
    ]

    found = []
    for text in texts:
        for pattern, label in COMMUNITY_AMENITY_PATTERNS:
            if re.search(pattern, text, flags=re.I):
                found.append(label)

    return unique_keep_order(found)


def extract_parking(main_page: Dict, amenities_page: Dict) -> List[str]:
    apartment_items, community_items, _ = extract_apartment_and_community_blocks(amenities_page)

    texts = [
        page_text(main_page),
        " ".join(apartment_items),
        " ".join(community_items),
        page_text(amenities_page),
    ]

    found = []
    for text in texts:
        for pattern, label in PARKING_PATTERNS:
            if re.search(pattern, text, flags=re.I):
                found.append(label)

    return unique_keep_order(found)


def extract_utilities(main_page: Dict, amenities_page: Dict) -> List[str]:
    _, _, utility_items = extract_apartment_and_community_blocks(amenities_page)

    found = []

    # Utilities section first
    utility_text = " ".join(utility_items)
    for pattern, label in INCLUDED_UTILITY_PATTERNS:
        if re.search(pattern, utility_text, flags=re.I):
            found.append(label)

    # Main description fallback
    post_content = extract_post_content(main_page)
    for pattern, label in INCLUDED_UTILITY_PATTERNS:
        if re.search(pattern, post_content, flags=re.I):
            found.append(label)

    return unique_keep_order(found)


def parse_property(prop: Dict) -> Dict[str, str]:
    base_url = clean(prop.get("base_url", ""))

    pages = prop.get("pages", {})
    main_page = pages.get("main", {})
    floorplans_page = pages.get("floorplans", {})
    amenities_page = pages.get("amenities", {})

    core = extract_core_fields(main_page, base_url)
    post_content = extract_post_content(main_page)
    suite_types = extract_suite_types(floorplans_page, main_page)
    suite_features = extract_suite_features(main_page, amenities_page)
    amenities = extract_community_amenities(main_page, amenities_page)
    parking = extract_parking(main_page, amenities_page)
    utilities = extract_utilities(main_page, amenities_page)

    return {
        "post_title": core["post_title"],
        "post_content": post_content,
        "property_type": "Apartment",
        "property_manager_name": "Realstar",
        "property_manager_phone": core["property_manager_phone"],
        "property_manager_website": core["property_manager_website"],
        "address": core["address"],
        "province": core["province"],
        "city_name": core["city_name"],
        "suite_types": "\n".join(suite_types),
        "suite_features": "\n".join(suite_features),
        "amenities": "\n".join(amenities),
        "utilities_included": "\n".join(utilities),
        "parking": "\n".join(parking),
    }


def write_outputs(rows: List[Dict[str, str]]) -> None:
    df = pd.DataFrame(rows, columns=HEADERS)
    df.to_csv(OUTPUT_CSV, index=False)

    template = Path(TEMPLATE_XLSX)
    if template.exists():
        wb = load_workbook(template)
        ws = wb["Import"]

        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        for r_idx, row in enumerate(rows, start=2):
            for c_idx, header in enumerate(HEADERS, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=row.get(header, ""))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[r_idx].height = 140

        wb.save(OUTPUT_XLSX)
    else:
        # plain XLSX fallback
        df.to_excel(OUTPUT_XLSX, index=False)


def main() -> None:
    raw = json.loads(Path(INPUT_JSON).read_text(encoding="utf-8"))
    rows = []

    for prop in raw:
        try:
            row = parse_property(prop)
            rows.append(row)
            print(f"Parsed: {row['post_title']}")
        except Exception as e:
            print(f"FAILED: {prop.get('base_url', '')} -> {e}")

    write_outputs(rows)
    print(f"Saved -> {OUTPUT_CSV}")
    print(f"Saved -> {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()